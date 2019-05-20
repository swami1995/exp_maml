# import cPickle
import numpy as np
import openpyxl as px
# import ipdb


def modify_excel(args, env_list=None):

	try:
		w = px.load_workbook('experiments_log.xlsx')
	except:
		create_workbook(args, env_list)
		return 0

	p = w.get_sheet_by_name('Experiments')
	# try:
	# 	p = w.get_sheet_by_name(args.env_name)
	# except:
	# 	create_worksheet(w, args)
	for i, row in enumerate(p.iter_rows()):
		if i==0:
			first_row = row
	arr = []
	# ipdb.set_trace()
	for c in first_row:
		arr.append(args[c.value])
	p.append(arr)
	w.save('experiments_log.xlsx')


def create_workbook(args, env_list):
	wb = px.Workbook()
	ws = wb.create_sheet('Experiments')
	# for env_name in env_list:
	# 	ws = wb.create_sheet(env_name)
	ws.append(list(args))
	ws.append(list(args.values()))
	wb.save('experiments_log.xlsx')
	print('created experiments_log.xlsx')

def create_worksheet(w, args):
	pass